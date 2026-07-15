import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# -------------------------------------------------------
# CHANGE THIS PATH
# -------------------------------------------------------
FOLDER_PATH = r"C:\Users\AlphaHP2\Desktop\DSA"

OUTPUT_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "DSA_Study_Tracker.xlsx",
)

VIDEO_EXTENSIONS = (
    ".mp4",
    ".mkv",
    ".avi",
    ".mov",
    ".wmv",
    ".flv",
    ".webm",
)


def sort_key(name):
    """Sort by lesson numbers like 1-1, 1-2, 2-1, 10-3."""
    match = re.match(r"^(\d+)-(\d+)", name)
    if match:
        return (int(match.group(1)), int(match.group(2)), name.lower())
    return (9999, 9999, name.lower())


# -------------------------------------------------------
# Create Workbook
# -------------------------------------------------------

wb = Workbook()
ws = wb.active
ws.title = "DSA Tracker"

headers = [
    "S.No",
    "Topic",
    "Status",
    "Start Date",
    "Done Date",
    "Revision 1",
    "Revision 2",
    "Practice Problems",
    "Notes",
    "Time Spent (Minutes)",
    "Confidence (1-5)",
]

header_fill = PatternFill(
    start_color="4F81BD",
    end_color="4F81BD",
    fill_type="solid",
)

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# -------------------------------------------------------
# Collect & sort videos (1-1, 1-2, ... 2-1, 2-2, ...)
# -------------------------------------------------------

videos = []

for root, dirs, files in os.walk(FOLDER_PATH):
    for file in files:
        if file.lower().endswith(VIDEO_EXTENSIONS):
            name = os.path.splitext(file)[0]
            videos.append(name)

videos.sort(key=sort_key)

row = 2
for count, name in enumerate(videos, start=1):
    ws.cell(row, 1).value = count
    ws.cell(row, 2).value = name
    ws.cell(row, 3).value = "☐"
    ws.cell(row, 4).value = ""
    ws.cell(row, 5).value = ""
    ws.cell(row, 6).value = ""
    ws.cell(row, 7).value = ""
    ws.cell(row, 8).value = 0
    ws.cell(row, 9).value = ""
    ws.cell(row, 10).value = 0
    ws.cell(row, 11).value = ""
    row += 1

# -------------------------------------------------------
# Auto Width
# -------------------------------------------------------

for column in ws.columns:
    length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
    ws.column_dimensions[get_column_letter(column[0].column)].width = min(
        length + 5, 40
    )

ws.freeze_panes = "A2"

wb.save(OUTPUT_FILE)

print(f"\nDone!")
print(f"Videos Found : {len(videos)}")
print(f"Excel Saved  : {OUTPUT_FILE}")
